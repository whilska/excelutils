from hilskapy.commons.logger import Logger, LoggerLevel
import argparse
import sys
import openpyxl as xl


class ExcelUtilities:

	def __init__(self, excel_path):
		self.log = Logger(self.__class__.__name__)
		self.excel_path = excel_path
		self.wb = None
		self.load_workbook(excel_path)

	def load_workbook(self, excel_path):
		self.wb = xl.load_workbook(excel_path)

	def get_excel_sheet_names(self):
		return self.wb.sheetnames

	def print_excel_sheet_names_to_log(self):
		sheet_names = self.get_excel_sheet_names()
		for i in sheet_names:
			self.log.info(i)


def main(args):
	Logger.set_logger_level(LoggerLevel.INFO)
	parser = argparse.ArgumentParser()
	parser.add_argument('excel_path', type=str, help='path to Excel file')
	excel_path = parser.parse_args(args).excel_path
	eu = ExcelUtilities(excel_path)
	eu.print_excel_sheet_names_to_log()


# args = [Excel path]
if __name__ == '__main__':
	main(sys.argv[1:])
